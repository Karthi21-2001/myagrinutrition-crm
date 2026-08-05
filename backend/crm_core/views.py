import csv
import json
import logging
import os
import openpyxl
import requests
import traceback

from datetime import timedelta
from functools import wraps

from django.contrib import messages
from django.contrib.auth import authenticate, get_user_model, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import AuthenticationForm
from django.core.serializers.json import DjangoJSONEncoder
from django.db import transaction
from django.db.models import Avg, Count, DecimalField, F, FloatField, Max, Q, Sum
from django.db.models.functions import Coalesce, TruncMonth, TruncYear
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .forms import ExecutiveSignUpForm
from .models import Farm, FarmVisitReport, VisitedProductDetail

logger = logging.getLogger(__name__)
User = get_user_model()


# ==========================================================
# DISTRICT NORMALIZATION MAP (backend mirror)
# ------------------------------------------------------------
# WHY THIS EXISTS: the reverse-geocode fix in the frontend form
# (farm_visit_form.html) only normalizes taluk/block/village names
# into real districts for NEW visits logged after that fix shipped.
# Farms saved before that still have raw taluk values sitting in
# Farm.district (Mohanur, Musiri, Rasipuram, Thottiyam, Tiruchengode,
# Kilvelur, Manachanallur, Agiripalle, Gannavaram, Nuzvid, Unguturu,
# etc). Those raw values were showing up as extra entries in the
# District filter dropdown alongside the real districts they belong
# to (Namakkal, Tiruchirappalli, ...).
#
# This mirrors the frontend's TALUK_TO_DISTRICT map. Keep both lists
# in sync when you add new mappings.
# ==========================================================
TALUK_TO_DISTRICT = {
    "mohanur": "Namakkal",
    "musiri": "Tiruchirappalli",
    "rasipuram": "Namakkal",
    "thottiyam": "Tiruchirappalli",
    "tiruchengode": "Namakkal",
    "kilvelur": "Nagapattinam",
    "manachanallur": "Tiruchirappalli",
    "agiripalle": "Krishna",
    "gannavaram": "NTR",
    "nuzvid": "Eluru",
    "unguturu": "Eluru",
    # Add more taluk/block/village -> district mappings as you spot
    # new noisy values appearing in the District filter dropdown.
}

# Reverse index: normalized district name (lowercased) -> every raw
# taluk/block value on record that should be treated as that district.
_DISTRICT_TO_RAW_TALUKS = {}
for _taluk, _district in TALUK_TO_DISTRICT.items():
    _DISTRICT_TO_RAW_TALUKS.setdefault(_district.lower(), set()).add(_taluk)


def normalize_district(raw_value):
    """Map a raw district/taluk/village string to its real district
    name. Unknown values are returned trimmed but otherwise untouched,
    so nothing gets silently dropped — it just won't be merged into a
    parent district until a mapping is added above.
    """
    if not raw_value:
        return ''
    key = raw_value.strip().lower()
    return TALUK_TO_DISTRICT.get(key, raw_value.strip())


def district_filter_values(selected_district):
    """Given a normalized district name picked from the dropdown
    (e.g. "Namakkal"), return every raw string already sitting in the
    database that should match it — the district name itself plus any
    taluk/block names that map to it (e.g. "Mohanur", "Rasipuram",
    "Tiruchengode"). Needed because older records were saved with the
    raw taluk value before the normalization fix existed.
    """
    key = selected_district.strip().lower()
    values = {key}
    values |= _DISTRICT_TO_RAW_TALUKS.get(key, set())
    return values


def build_district_q(field_path, selected_district):
    """Build a Q object matching any raw value (district or taluk)
    that normalizes to `selected_district`, for the given ORM field
    lookup path (e.g. "district", "farm__district",
    "visit__farm__district").
    """
    q = Q()
    for val in district_filter_values(selected_district):
        q |= Q(**{f"{field_path}__iexact": val})
    return q


def staff_required(view_func):
    """Restrict a view to staff/superuser (admin) accounts only.

    Field executives are NOT staff — this is what actually keeps them
    off the dashboard, analytics, and Excel export pages even if they
    type the URL directly or bookmark it (login_required alone isn't
    enough, since it only checks that *someone* is logged in).

    - Anonymous users are sent to the login page.
    - Authenticated non-staff users (executives) are sent back to
      their visit-logging form, not an access-denied page — from
      their point of view the CRM only ever has one screen.
    - Staff/superusers pass through untouched.
    """
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login_user')
        if not (request.user.is_staff or request.user.is_superuser):
            return redirect('render_visit_form')
        return view_func(request, *args, **kwargs)
    return _wrapped


# ==========================================
# 🔐 EXECUTIVE AUTHENTICATION CONTROLLERS
# ==========================================

def register_user(request):
    if request.method == 'POST':
        form = ExecutiveSignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            if user.is_staff or user.is_superuser:
                return redirect('dashboard_home')
            return redirect('render_visit_form')
    else:
        form = ExecutiveSignUpForm()
    return render(request, 'crm_core/register.html', {'form': form})


def login_user(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                if user.is_staff or user.is_superuser:
                    return redirect('dashboard_home')
                return redirect('render_visit_form')
    else:
        form = AuthenticationForm()
    return render(request, 'crm_core/login.html', {'form': form})


def logout_user(request):
    logout(request)
    return redirect('login_user')


# ==========================================
# 🌱 AGRI-CORE MANAGEMENT FUNCTIONALITY
# ==========================================

@login_required(login_url='/crm/login/')
def render_visit_form(request):
    return render(request, 'crm_core/farm_visit_form.html')


def _to_int(val):
    """Safely coerce a POSTed form value to int, defaulting to 0."""
    try:
        return int(val)
    except (TypeError, ValueError):
        return 0


def _to_float(val):
    """Safely coerce a POSTed form value to float, defaulting to 0.0."""
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _to_date_or_none(val):
    """Safely coerce a POSTed HTML <input type="date"> value
    (expected format YYYY-MM-DD) to a date object, or None if blank
    or unparsable. Kept permissive here — Django's DateField will do
    the authoritative validation/casting on save.
    """
    val = (val or '').strip()
    if not val:
        return None
    return val


@login_required(login_url='/crm/login/')
def save_farm_visit(request):
    if request.method == 'POST':
        farm_name = request.POST.get('farm_name')
        owner_name = request.POST.get('owner_name')
        contact_number = request.POST.get('contact_number')
        business_type = request.POST.get('business_type', 'Poultry')

        # FIX: the form field's `name` attribute is `sub_business_type`
        # (the old key `sub_business_type_select` was actually the
        # element's `id`, not its POST name, so this always returned
        # '' and Sub-Segment was saved blank for every visit).
        sub_segment = request.POST.get('sub_business_type', '').strip()

        # Normalize here too as a backend safety net (the reverse-geocode
        # widget already normalizes taluk/block names client-side, but
        # this covers direct POSTs and keeps behavior correct either way).
        district = normalize_district(request.POST.get('district', ''))
        area = request.POST.get('area', '').strip()
        state = request.POST.get('state', '').strip()
        farm_problem = request.POST.get('farm_problem')

        # Planned follow-up date captured on the "Next Visit Date"
        # field in the form. Stored on the visit record it was logged
        # from, since it's a per-visit follow-up plan rather than a
        # permanent attribute of the farm itself.
        next_visit_date = _to_date_or_none(request.POST.get('next_visit_date'))

        if not state or state.lower() in ['state', 'unknown state', '']:
            state = 'Tamil Nadu'

        lat = request.POST.get('latitude')
        lon = request.POST.get('longitude')
        latitude = float(lat) if lat else None
        longitude = float(lon) if lon else None

        # FIX: these Poultry Shed Population Inventory fields were being
        # submitted by the form (chicks_count, grower_count, layer_count,
        # culling_bird_count) but were never read from request.POST or
        # saved onto the Farm record anywhere in this view, so every
        # visit was persisted with counts stuck at 0.
        chicks_count = _to_int(request.POST.get('chicks_count'))
        grower_count = _to_int(request.POST.get('grower_count'))
        layer_count = _to_int(request.POST.get('layer_count'))
        culling_bird_count = _to_int(request.POST.get('culling_bird_count'))

        # Aqua Pond Tracking Inventory fields (captured for completeness;
        # only persisted if the Farm model has matching fields).
        pond_acre = _to_float(request.POST.get('pond_acre'))
        pond_doc = _to_int(request.POST.get('pond_doc'))
        fish_variety = request.POST.get('fish_variety', '').strip()

        current_user = request.user if request.user.is_authenticated else None

        try:
            with transaction.atomic():
                farm_instance, created = Farm.objects.get_or_create(
                    farm_name=farm_name,
                    owner_name=owner_name,
                    contact_number=contact_number,
                    defaults={
                        'executive': current_user,
                        'business_type': business_type,
                        'sub_segment': sub_segment,
                        'state': state,
                        'district': district,
                        'area': area,
                        'latitude': latitude,
                        'longitude': longitude,
                        'chicks_count': chicks_count,
                        'grower_count': grower_count,
                        'layer_count': layer_count,
                        'culling_bird_count': culling_bird_count,
                    }
                )

                if not created:
                    if business_type:
                        farm_instance.business_type = business_type
                    if sub_segment:
                        farm_instance.sub_segment = sub_segment
                    farm_instance.state = state
                    farm_instance.district = district
                    farm_instance.area = area
                    farm_instance.chicks_count = chicks_count
                    farm_instance.grower_count = grower_count
                    farm_instance.layer_count = layer_count
                    farm_instance.culling_bird_count = culling_bird_count
                    farm_instance.save()

                # NOTE: requires a `next_visit_date = models.DateField(null=True, blank=True)`
                # field on the FarmVisitReport model — add it there if it
                # doesn't already exist, then makemigrations/migrate.
                # Guarded with a fallback so a not-yet-migrated field can't
                # 500 the whole visit save — it just gets skipped that time.
                try:
                    visit_record = FarmVisitReport.objects.create(
                        farm=farm_instance,
                        executive=current_user,
                        farm_problem=farm_problem,
                        next_visit_date=next_visit_date
                    )
                except TypeError:
                    logger.warning(
                        "FarmVisitReport has no next_visit_date field yet — "
                        "add it to models.py and run migrations. Saving visit without it."
                    )
                    visit_record = FarmVisitReport.objects.create(
                        farm=farm_instance,
                        executive=current_user,
                        farm_problem=farm_problem
                    )

                # Process Sales Order Products
                order_products = request.POST.getlist('discussed_product[]')
                sale_quantities = request.POST.getlist('sale_quantity[]')
                unit_types = request.POST.getlist('unit_type[]')
                primary_prices = request.POST.getlist('primary_price[]')

                for i in range(len(order_products)):
                    prod_name = order_products[i].strip()
                    if not prod_name:
                        continue

                    s_qty = int(sale_quantities[i]) if (i < len(sale_quantities) and sale_quantities[i]) else 0
                    unit = unit_types[i] if i < len(unit_types) else 'KG'
                    price = float(primary_prices[i]) if (i < len(primary_prices) and primary_prices[i]) else 0.00

                    VisitedProductDetail.objects.create(
                        visit=visit_record,
                        product_name=prod_name,
                        potential_quantity=0,
                        target_quantity=0,
                        sale_quantity=s_qty,
                        unit_type=unit,
                        primary_price=price,
                        revenue_generated=price * s_qty,
                        process_status='Hot' if s_qty > 0 else 'Warm',
                        conversion_percentage=100 if s_qty > 0 else 0
                    )

                # Process Pipeline Products
                pipeline_products = request.POST.getlist('pipeline_discussed_product[]')
                p_quantities = request.POST.getlist('pipeline_potential_quantity[]')
                t_quantities = request.POST.getlist('pipeline_target_quantity[]')
                p_unit_types = request.POST.getlist('pipeline_unit_type[]')
                p_statuses = request.POST.getlist('pipeline_process_status[]')
                p_conv_percentages = request.POST.getlist('pipeline_conversion_percentage[]')

                for i in range(len(pipeline_products)):
                    pipe_prod_name = pipeline_products[i].strip()
                    if not pipe_prod_name:
                        continue

                    p_qty = int(p_quantities[i]) if (i < len(p_quantities) and p_quantities[i]) else 0
                    t_qty = int(t_quantities[i]) if (i < len(t_quantities) and t_quantities[i]) else 0
                    p_unit = p_unit_types[i] if i < len(p_unit_types) else 'KG'
                    status = p_statuses[i] if i < len(p_statuses) else 'Warm'
                    conv_pct = int(p_conv_percentages[i]) if (i < len(p_conv_percentages) and p_conv_percentages[i]) else 0

                    VisitedProductDetail.objects.create(
                        visit=visit_record,
                        product_name=pipe_prod_name,
                        potential_quantity=p_qty,
                        target_quantity=t_qty,
                        sale_quantity=0,
                        primary_price=0.00,
                        revenue_generated=0.00,
                        unit_type=p_unit,
                        process_status=status,
                        conversion_percentage=conv_pct
                    )

            messages.success(request, "Agri-Field visit logging record processed successfully!")
            if request.user.is_staff or request.user.is_superuser:
                return redirect('dashboard_home')

            return render(request, 'crm_core/farm_visit_form.html', {'saved_data': request.POST})

        except Exception as e:
            logger.error(f"Error in save_farm_visit: {str(e)}", exc_info=True)
            messages.error(request, f"Database transaction block failed: {str(e)}")
            return render(request, 'crm_core/farm_visit_form.html', {'saved_data': request.POST})

    return redirect('render_visit_form')


# ==========================================
# 📥 EXCEL EXPORT ENGINE
# ==========================================

@staff_required
@csrf_exempt
def export_visits_to_excel(request):
    start_date_str = request.GET.get('start_date', '').strip()
    end_date_str = request.GET.get('end_date', '').strip()
    executive_filter = request.GET.get('executive', '').strip()
    state_filter = request.GET.get('state', '').strip()
    district_filter = request.GET.get('district', '').strip()
    business_type = request.GET.get('business_type', '').strip()
    sub_segment_filter = request.GET.get('sub_segment', '').strip()

    export_filters = Q()

    if business_type and business_type not in ['All', 'All Sectors']:
        export_filters &= Q(farm__business_type__iexact=business_type)
    if sub_segment_filter and sub_segment_filter != 'All':
        export_filters &= Q(farm__sub_segment__iexact=sub_segment_filter)
    if state_filter and state_filter not in ['All', 'All States']:
        export_filters &= Q(farm__state__iexact=state_filter)
    if district_filter and district_filter not in ['All', 'All Districts']:
        export_filters &= build_district_q('farm__district', district_filter)
    if executive_filter and executive_filter not in ['All', 'All Executives']:
        export_filters &= Q(executive__username__iexact=executive_filter)

    if start_date_str:
        try:
            export_filters &= Q(visit_date__date__gte=start_date_str)
        except ValueError:
            pass
    if end_date_str:
        try:
            export_filters &= Q(visit_date__date__lte=end_date_str)
        except ValueError:
            pass

    wb = openpyxl.Workbook()
    ws_data = wb.active
    ws_data.title = "Field Visit Database Log"
    ws_data.views.sheetView[0].showGridLines = True

    dark_slate, border_color = "0F172A", "CBD5E1"
    thin_border = Border(
        left=Side(style='thin', color=border_color), right=Side(style='thin', color=border_color),
        top=Side(style='thin', color=border_color), bottom=Side(style='thin', color=border_color)
    )

    headers = [
        'Visit Date', 'Next Visit Date', 'Executive Name', 'Farm Name', 'Owner Name', 'Contact Number',
        'Sector Segment', 'Sub-Segment', 'State', 'District', 'Area / Suburb',
        'Farm Problem Observed', 'Chicks Count', 'Grower Count', 'Layer Count', 'Culling Bird',
        'Product Name', 'Sale Qty', 'Price (INR)', 'Revenue Generated',
        'Poten. Qty', 'Target Qty', 'Units', 'Process Stage', 'conv (%)', 'Live GPS Link'
    ]

    for col_idx, text in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=text)
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=dark_slate, end_color=dark_slate, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_data.row_dimensions[1].height = 28

    all_visits = FarmVisitReport.objects.filter(export_filters).select_related(
        'farm', 'executive'
    ).order_by('-visit_date')

    current_row = 2
    for v in all_visits:
        f = v.farm if v else None
        products = VisitedProductDetail.objects.filter(visit=v)
        product_loop_list = products if products.exists() else [None]

        for p in product_loop_list:
            # FIX (IST bug): visit_date is stored in UTC (USE_TZ=True).
            # Calling .strftime() directly on it printed the raw UTC
            # value instead of converting to the local TIME_ZONE
            # ('Asia/Kolkata'). timezone.localtime() does that
            # conversion, which is what template date filters do
            # automatically but raw Python code must do explicitly.
            ws_data.cell(
                row=current_row,
                column=1,
                value=timezone.localtime(v.visit_date).strftime("%Y-%m-%d %H:%M") if v and v.visit_date else ""
            )

            # FIX (Next Visit Date blank in export): this was already
            # correctly reading getattr(v, 'next_visit_date', None) —
            # the blank cells were caused by the field not existing on
            # FarmVisitReport yet (see save_farm_visit's try/except
            # TypeError fallback), not by this line. Once
            # next_visit_date is added to models.py and migrated, new
            # visits will populate here automatically.
            nvd = getattr(v, 'next_visit_date', None) if v else None
            ws_data.cell(row=current_row, column=2, value=nvd.strftime("%Y-%m-%d") if nvd else "")

            ws_data.cell(row=current_row, column=3, value=v.executive.username if v and v.executive else "")
            ws_data.cell(row=current_row, column=4, value=f.farm_name if f else "")
            ws_data.cell(row=current_row, column=5, value=f.owner_name if f else "")
            ws_data.cell(row=current_row, column=6, value=f.contact_number if f else "")
            ws_data.cell(row=current_row, column=7, value=f.business_type if f else "")
            ws_data.cell(row=current_row, column=8, value=f.sub_segment if (f and f.sub_segment) else "")
            ws_data.cell(row=current_row, column=9, value=f.state if f else "")
            ws_data.cell(row=current_row, column=10, value=normalize_district(f.district) if f else "")
            ws_data.cell(row=current_row, column=11, value=f.area if f else "")

            ws_data.cell(row=current_row, column=12, value=v.farm_problem if (v and v.farm_problem) else "None reported")

            ws_data.cell(row=current_row, column=13, value=getattr(f, 'chicks_count', 0) if f else 0)
            ws_data.cell(row=current_row, column=14, value=getattr(f, 'grower_count', 0) if f else 0)
            ws_data.cell(row=current_row, column=15, value=getattr(f, 'layer_count', 0) if f else 0)
            ws_data.cell(row=current_row, column=16, value=getattr(f, 'culling_bird_count', 0) if f else 0)

            ws_data.cell(row=current_row, column=17, value=p.product_name if p else "General Consult")
            ws_data.cell(row=current_row, column=18, value=p.sale_quantity if p else 0)
            ws_data.cell(row=current_row, column=19, value=float(p.primary_price) if p else 0.0)
            ws_data.cell(row=current_row, column=20, value=float(p.revenue_generated) if p else 0.0)

            ws_data.cell(row=current_row, column=21, value=p.potential_quantity if p else 0)
            ws_data.cell(row=current_row, column=22, value=p.target_quantity if p else 0)
            ws_data.cell(row=current_row, column=23, value=p.unit_type if p else "N/A")
            ws_data.cell(row=current_row, column=24, value=p.process_status if p else "N/A")
            ws_data.cell(row=current_row, column=25, value=f"{p.conversion_percentage}%" if p else "0%")

            gps_cell = ws_data.cell(row=current_row, column=26)
            if f and f.latitude and f.longitude:
                gps_cell.value = "View on Map"
                gps_cell.hyperlink = f"https://maps.google.com/?q={f.latitude},{f.longitude}"
                gps_cell.font = Font(name="Segoe UI", size=11, color="0000FF", underline="single")
            else:
                gps_cell.value = "No GPS Data"
                gps_cell.font = Font(name="Segoe UI", size=11, color="64748B", italic=True)

            for c_idx in range(1, 27):
                cell_item = ws_data.cell(row=current_row, column=c_idx)
                cell_item.border = thin_border
                if c_idx in [13, 14, 15, 16, 18, 21, 22, 25]:
                    cell_item.alignment = Alignment(horizontal="center")

            current_row += 1

    if ws_data.max_row > 1:
        for col in ws_data.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_data.column_dimensions[col_letter].width = max(max_len + 4, 14)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="MyAgrinutrition_CRM_Field_Logs.xlsx"'
    wb.save(response)
    return response


# ==========================================
# 📊 DASHBOARDS & ADVANCED ANALYTICS PIPELINES
# ==========================================


def get_dashboard_context(request):
    """Computes all KPI aggregates, Chart data series, and list structures

    for the MyAgriNutrition Analytics Dashboard safely with complete error protection.
    """
    # ------------------------------------------------------------------
    # 1. READ FILTER INPUTS FROM REQUEST
    # ------------------------------------------------------------------
    sel_state = request.GET.get("state", "").strip()
    sel_country = request.GET.get("country", "").strip()
    sel_district = request.GET.get("district", "").strip()
    sel_executive = request.GET.get("executive", "").strip()
    sel_month = request.GET.get("month", "").strip()
    sel_year = request.GET.get("year", "").strip()
    sel_sector = request.GET.get("sector", "").strip()
    sel_start_date = request.GET.get("start_date", "").strip()
    sel_end_date = request.GET.get("end_date", "").strip()

    # ------------------------------------------------------------------
    # 2. INITIALIZE SAFE DEFAULT CONTEXT
    # ------------------------------------------------------------------
    context = {
        # Numeric KPIs
        "total_revenue": 0.0,
        "total_visits": 0,
        "active_executives": 0,
        "total_farms": 0,
        "total_sales_volume": 0,
        "paid_orders_count": 0,
        "avg_order_value": 0.0,
        "conversion_rate": 0.0,
        # Percentages
        "hot_pct": 0.0,
        "warm_pct": 0.0,
        "cold_pct": 0.0,
        "poultry_pct": 0.0,
        "aqua_pct": 0.0,
        # Environment flag — controls whether the WhatsApp Notify button
        # renders at all in dashboard.html. Selenium/Chrome only work on
        # a local machine with a display and a logged-in WhatsApp Web
        # session, never on Render, so the button is hidden entirely
        # there instead of showing an always-fails error popup. Render
        # automatically sets RENDER=true on every deployment, so this
        # needs no extra configuration on your end.
        "is_local_environment": os.environ.get("IS_LOCAL_DEV", "false").lower() == "true",
        # Pre-Serialized Safe JSON for Chart.js
        "month_wise_labels_js": json.dumps([]),
        "month_wise_data_js": json.dumps([]),
        "year_wise_labels_js": json.dumps([]),
        "year_wise_data_js": json.dumps([]),
        "exec_labels_js": json.dumps([]),
        "exec_revenue_js": json.dumps([]),
        "exec_conv_pct_js": json.dumps([]),
        "top_prod_labels_js": json.dumps([]),
        "top_prod_revenue_js": json.dumps([]),
        "top_prod_qty_js": json.dumps([]),
        "state_labels_js": json.dumps([]),
        "state_data_js": json.dumps([]),
        "chart_labels_js": [],
        "chart_counts_js": [],
        "farm_locations_js": [],
        "stale_farms": [],
        "stale_farms_count": 0,
        "prob_labels_js": json.dumps([]),
        "prob_data_js": json.dumps([]),
        "bird_labels_js": json.dumps(
            ["Chicks", "Growers", "Layers", "Culling Birds"]
        ),
        "bird_counts_js": json.dumps([0, 0, 0, 0]),
        # Structured Tables & Querysets
        "pipeline_spread": {"actual": 0, "target": 0, "potential": 0},
        "top_farms": [],
        "recent_visits": [],
        # Options
        "state_list": [],
        "district_list": [],
        "executive_list": [],
        "year_list": [],
        "country_list": ["India"],
        # Retention
        "selected_state": sel_state,
        "selected_country": sel_country,
        "selected_district": sel_district,
        "selected_executive": sel_executive,
        "selected_month": sel_month,
        "selected_year": sel_year,
        "selected_sector": sel_sector,
        "start_date": sel_start_date,
        "end_date": sel_end_date,
    }

    try:
        # ------------------------------------------------------------------
        # 3. CONSTRUCT FILTER CONDITIONS
        # ------------------------------------------------------------------
        farm_filters = Q()
        visit_filters = Q()
        product_filters = Q()

        if sel_state and sel_state not in ["All", "All States", ""]:
            farm_filters &= Q(state__iexact=sel_state)
            visit_filters &= Q(farm__state__iexact=sel_state)
            product_filters &= Q(visit__farm__state__iexact=sel_state)

        if sel_district and sel_district not in ["All", "All Districts", ""]:
            farm_filters &= build_district_q('district', sel_district)
            visit_filters &= build_district_q('farm__district', sel_district)
            product_filters &= build_district_q('visit__farm__district', sel_district)

        if sel_executive and sel_executive not in ["All", "All Executives", ""]:
            farm_filters &= Q(executive__username__iexact=sel_executive)
            visit_filters &= Q(executive__username__iexact=sel_executive)
            product_filters &= Q(
                visit__executive__username__iexact=sel_executive
            )

        if sel_month and sel_month not in ["All", "All Months", ""]:
            try:
                m_val = int(sel_month)
                visit_filters &= Q(visit_date__month=m_val)
                product_filters &= Q(visit__visit_date__month=m_val)
            except ValueError:
                pass

        if sel_year and sel_year not in ["All", ""]:
            try:
                y_val = int(sel_year)
                visit_filters &= Q(visit_date__year=y_val)
                product_filters &= Q(visit__visit_date__year=y_val)
            except ValueError:
                pass

        if sel_sector and sel_sector not in ["All", "All Sectors", ""]:
            farm_filters &= Q(business_type__icontains=sel_sector)
            visit_filters &= Q(farm__business_type__icontains=sel_sector)
            product_filters &= Q(
                visit__farm__business_type__icontains=sel_sector
            )

        if sel_start_date:
            try:
                visit_filters &= Q(visit_date__date__gte=sel_start_date)
                product_filters &= Q(
                    visit__visit_date__date__gte=sel_start_date
                )
            except ValueError:
                pass

        if sel_end_date:
            try:
                visit_filters &= Q(visit_date__date__lte=sel_end_date)
                product_filters &= Q(
                    visit__visit_date__date__lte=sel_end_date
                )
            except ValueError:
                pass

        # Base Querysets
        visit_qs = FarmVisitReport.objects.filter(visit_filters)
        farm_qs = Farm.objects.filter(farm_filters)
        product_qs = VisitedProductDetail.objects.filter(product_filters)

        # ------------------------------------------------------------------
        # 4. PRIMARY METRICS ACCUMULATION
        # ------------------------------------------------------------------
        v_count = visit_qs.count()
        total_farms_count = farm_qs.count()

        active_execs_qs = (
            visit_qs.exclude(
                Q(executive__isnull=True) | Q(executive__username="")
            )
            .values("executive")
            .distinct()
        )
        active_executives = active_execs_qs.count()

        total_rev = round(float(
            product_qs.aggregate(total=Coalesce(Sum("revenue_generated"), 0.0, output_field=FloatField()))[
                "total"
            ]
        ), 2)
        vol_sold = int(
            product_qs.aggregate(total_qty=Coalesce(Sum("sale_quantity"), 0))[
                "total_qty"
            ]
        )

        paid_orders_count = product_qs.filter(
            Q(sale_quantity__gt=0) | Q(revenue_generated__gt=0)
        ).count()

        avg_order_value = (
            round(float(total_rev / paid_orders_count), 2)
            if paid_orders_count > 0
            else 0.0
        )

        poultry_pct, aqua_pct = 0.0, 0.0
        if total_farms_count > 0:
            p_cnt = farm_qs.filter(business_type__icontains="Poultry").count()
            a_cnt = farm_qs.filter(business_type__icontains="Aqua").count()
            poultry_pct = round((p_cnt / total_farms_count) * 100, 1)
            aqua_pct = round((a_cnt / total_farms_count) * 100, 1)

        total_leads = product_qs.count()
        hot_pct, warm_pct, cold_pct = 0.0, 0.0, 0.0
        if total_leads > 0:
            h_cnt = product_qs.filter(process_status__iexact="Hot").count()
            w_cnt = product_qs.filter(process_status__iexact="Warm").count()
            c_cnt = product_qs.filter(process_status__iexact="Cold").count()

            hot_pct = round((h_cnt / total_leads) * 100, 1)
            warm_pct = round((w_cnt / total_leads) * 100, 1)
            cold_pct = round((c_cnt / total_leads) * 100, 1)

        conversion_rate = hot_pct

        # ------------------------------------------------------------------
        # 5. CHART DATA SERIES GENERATION
        # ------------------------------------------------------------------
        month_wise_qs = list(
            product_qs.annotate(month=TruncMonth("visit__visit_date"))
            .values("month")
            .annotate(revenue=Coalesce(Sum("revenue_generated"), 0.0, output_field=FloatField()))
            .filter(month__isnull=False)
            .order_by("month")
        )
        month_wise_labels = [
            m["month"].strftime("%b %Y")
            for m in month_wise_qs
            if m.get("month")
        ]
        month_wise_data = [float(m["revenue"]) for m in month_wise_qs]

        year_wise_qs = list(
            product_qs.annotate(year=TruncYear("visit__visit_date"))
            .values("year")
            .annotate(revenue=Coalesce(Sum("revenue_generated"), 0.0, output_field=FloatField()))
            .filter(year__isnull=False)
            .order_by("year")
        )
        year_wise_labels = [
            y["year"].strftime("%Y") for y in year_wise_qs if y.get("year")
        ]
        year_wise_data = [float(y["revenue"]) for y in year_wise_qs]

        exec_perf = (
            product_qs.values("visit__executive__username")
            .annotate(
                revenue=Coalesce(Sum("revenue_generated"), 0.0, output_field=FloatField()),
                total_items=Count("id"),
                hot_items=Count("id", filter=Q(process_status__iexact="Hot")),
            )
            .order_by("-revenue")[:10]
        )
        exec_labels = [
            e["visit__executive__username"] or "Unassigned" for e in exec_perf
        ]
        exec_revenue = [float(e["revenue"]) for e in exec_perf]
        exec_conv_pct = [
            round((e["hot_items"] / e["total_items"] * 100), 1)
            if e["total_items"] > 0
            else 0.0
            for e in exec_perf
        ]

        prod_perf = (
            product_qs.values("product_name")
            .annotate(
                revenue=Coalesce(Sum("revenue_generated"), 0.0, output_field=FloatField()),
                qty_sold=Coalesce(Sum("sale_quantity"), 0),
            )
            .exclude(Q(product_name__isnull=True) | Q(product_name=""))
            .order_by("-qty_sold")[:6]
        )
        top_prod_labels = [p["product_name"] for p in prod_perf]
        top_prod_revenue = [float(p["revenue"]) for p in prod_perf]
        top_prod_qty = [int(p["qty_sold"]) for p in prod_perf]

        state_qs = (
            visit_qs.values("farm__state")
            .annotate(total_visits=Count("id"))
            .exclude(Q(farm__state__isnull=True) | Q(farm__state=""))
            .order_by("-total_visits")[:6]
        )
        state_labels = [s["farm__state"] for s in state_qs]
        state_data = [s["total_visits"] for s in state_qs]

        district_qs = (
            farm_qs.values("district")
            .annotate(farm_count=Count("id"))
            .exclude(Q(district__isnull=True) | Q(district=""))
        )
        # Normalize raw taluk/village values into their parent district
        # before aggregating, so e.g. "Mohanur" and "Rasipuram" counts
        # merge into the "Namakkal" bar instead of showing separately.
        district_counts = {}
        for d in district_qs:
            normalized = normalize_district(d["district"])
            district_counts[normalized] = district_counts.get(normalized, 0) + d["farm_count"]
        sorted_districts = sorted(district_counts.items(), key=lambda kv: kv[1], reverse=True)[:10]
        chart_labels = [name for name, _ in sorted_districts]
        chart_counts = [count for _, count in sorted_districts]

        # ------------------------------------------------------------------
        # 5b. FARM COVERAGE MAP DATA
        # ------------------------------------------------------------------
        # Feeds the Leaflet coverage map on dashboard.html — one entry per
        # farm that has GPS coordinates on file (captured at visit-logging
        # time in save_farm_visit / the reverse-geocode widget). Farms
        # without lat/lng are skipped here; the template counts and warns
        # about them separately via a banner rather than silently omitting
        # them.
        #
        # NOTE: this is built from visit_qs (not farm_qs). farm_qs only
        # ever applies state/district/executive/sector — it never gained
        # month/year/start_date/end_date filtering, so date-range picks
        # from the dashboard toolbar silently had no effect on the map.
        # visit_qs already has every filter applied (including the date
        # range), so pulling the distinct set of farms referenced by
        # those visits makes the map respect the same filters as the
        # rest of the dashboard. Every Farm row is created alongside its
        # first FarmVisitReport (see save_farm_visit's get_or_create), so
        # this is equivalent to farm_qs whenever no date/month/year
        # filter narrows visit_qs below farm_qs.
        map_farm_ids = visit_qs.values_list("farm_id", flat=True).distinct()
        farm_locations = list(
            Farm.objects.filter(id__in=map_farm_ids)
            .exclude(Q(latitude__isnull=True) | Q(longitude__isnull=True))
            .values(
                "id",
                "farm_name",
                "owner_name",
                "district",
                "state",
                "business_type",
                "latitude",
                "longitude",
            )
        )

        # ------------------------------------------------------------------
        # 5c. VISIT STALENESS ("coverage gap") CALCULATION
        # ------------------------------------------------------------------
        # This is deliberately computed from ALL of a farm's visit history
        # (not visit_qs / not scoped to month/year/date-range filters) —
        # "days since last visit" is a real-world fact about the farm, and
        # scoping it to the currently selected date range would make a
        # farm visited yesterday show as "stale" just because the person
        # happens to be looking at last month's data. State/district/
        # executive/sector filters still apply (via farm_qs) since those
        # narrow *which farms* are in view, not *when* they're compared.
        STALE_THRESHOLD_DAYS = 30
        now = timezone.now()

        last_visit_map = {
            row["farm_id"]: row["last_visit"]
            for row in FarmVisitReport.objects.values("farm_id").annotate(
                last_visit=Max("visit_date")
            )
        }

        def _days_since(farm_id):
            last_visit = last_visit_map.get(farm_id)
            if not last_visit:
                return None  # never visited
            return (now - last_visit).days

        farm_locations_data = []
        for f in farm_locations:
            days_since = _days_since(f["id"])
            farm_locations_data.append(
                {
                    "name": f["farm_name"],
                    "owner": f["owner_name"],
                    "district": normalize_district(f["district"]),
                    "state": f["state"],
                    "business_type": f["business_type"],
                    "lat": float(f["latitude"]),
                    "lng": float(f["longitude"]),
                    "stale": days_since is None or days_since >= STALE_THRESHOLD_DAYS,
                    "days_since_visit": days_since,
                }
            )

        # Standalone "coverage gap" list — every farm in the current
        # state/district/executive/sector scope (farm_qs) that hasn't
        # been visited in STALE_THRESHOLD_DAYS+ days, or has never been
        # visited at all. Not restricted to farms with GPS coordinates,
        # since this list is meant to drive follow-up action even before
        # a farm has a pin on the map.
        stale_farms_raw = []
        for f in farm_qs.values(
            "id", "farm_name", "owner_name", "district", "state",
            "business_type", "executive__username",
        ):
            days_since = _days_since(f["id"])
            if days_since is None or days_since >= STALE_THRESHOLD_DAYS:
                stale_farms_raw.append(
                    {
                        "name": f["farm_name"],
                        "owner": f["owner_name"],
                        "district": normalize_district(f["district"]),
                        "state": f["state"],
                        "business_type": f["business_type"],
                        "executive": f["executive__username"] or "Unassigned",
                        "days_since_visit": days_since,
                    }
                )

        # Sort worst-first: never-visited farms (None) float to the top,
        # then longest-overdue first.
        #
        # FIX: the old key sorted the tie-breaker (days_since_visit)
        # ascending, so among visited farms the *least* overdue ones
        # came first and the most overdue ones got pushed toward the
        # bottom — the opposite of "longest-overdue first", and they
        # could even fall outside the [:25] slice below. Negating the
        # day count flips that tier to descending while keeping the
        # never-visited farms in front.
        stale_farms_raw.sort(
            key=lambda x: (
                x["days_since_visit"] is not None,  # None (never-visited) sorts first
                -(x["days_since_visit"] or 0),       # then most-overdue days first
            )
        )
        stale_farms_count = len(stale_farms_raw)
        stale_farms_table = stale_farms_raw[:25]

        prob_qs = (
            visit_qs.values("farm_problem")
            .annotate(frequency=Count("id"))
            .exclude(Q(farm_problem__isnull=True) | Q(farm_problem=""))
            .order_by("-frequency")[:5]
        )
        prob_labels = [p["farm_problem"] for p in prob_qs]
        prob_data = [p["frequency"] for p in prob_qs]

        # ------------------------------------------------------------------
        # 6. DEMOGRAPHICS & LIST DATA
        # ------------------------------------------------------------------
        top_farms_raw = (
            product_qs.values(
                "visit__farm__farm_name", "visit__farm__owner_name"
            )
            .annotate(revenue=Coalesce(Sum("revenue_generated"), 0.0, output_field=FloatField()))
            .filter(revenue__gt=0)
            .order_by("-revenue")[:5]
        )
        # Re-keyed to plain "name"/"owner"/"revenue" so templates can use
        # {{ farm.name }} instead of the raw dunder lookup path.
        top_farms_table = [
            {
                "name": f["visit__farm__farm_name"],
                "owner": f["visit__farm__owner_name"],
                "revenue": f["revenue"],
            }
            for f in top_farms_raw
        ]

        try:
            bird_population = farm_qs.aggregate(
                chicks=Coalesce(Sum("chicks_count"), 0),
                growers=Coalesce(Sum("grower_count"), 0),
                layers=Coalesce(Sum("layer_count"), 0),
                culling=Coalesce(Sum("culling_bird_count"), 0),
            )
            bird_counts = [
                bird_population["chicks"],
                bird_population["growers"],
                bird_population["layers"],
                bird_population["culling"],
            ]
        except Exception as e:
            logger.warning(f"Bird count aggregation skipped: {e}")
            bird_counts = [0, 0, 0, 0]

        pipeline_spread_agg = product_qs.aggregate(
            actual=Coalesce(Sum("sale_quantity"), 0),
            target=Coalesce(Sum("target_quantity"), 0),
            potential=Coalesce(Sum("potential_quantity"), 0),
        )

        recent_visits = list(
            visit_qs.select_related("farm", "executive").order_by(
                "-visit_date"
            )[:10]
        )
        # analytics_report.html reads visit.calculated_total, which isn't a
        # model field — attach it here as the visit's total product revenue.
        for rv in recent_visits:
            rv.calculated_total = float(
                VisitedProductDetail.objects.filter(visit=rv).aggregate(
                    total=Coalesce(Sum("revenue_generated"), 0.0, output_field=FloatField())
                )["total"]
            )

        # --------------------------------------------------------------
        # Dedup fix: .distinct() only removes EXACT string matches, so
        # values differing only by casing/whitespace (e.g. "Tamil Nadu"
        # vs "tamil nadu ") were showing up as separate dropdown entries.
        # We now normalize (strip + lowercase) for comparison while
        # keeping the first-seen original casing for display.
        # --------------------------------------------------------------
        raw_states = Farm.objects.exclude(
            Q(state__isnull=True) | Q(state="")
        ).values_list("state", flat=True)
        seen_states = {}
        for s in raw_states:
            key = s.strip().lower()
            if key and key not in seen_states:
                seen_states[key] = s.strip()
        state_list = sorted(seen_states.values())

        raw_districts = Farm.objects.exclude(
            Q(district__isnull=True) | Q(district="")
        ).values_list("district", flat=True)
        seen_districts = {}
        for d in raw_districts:
            # Normalize taluk/block/village values (e.g. "Mohanur",
            # "Rasipuram") into their real district ("Namakkal") before
            # dedup, so the dropdown only ever shows real district names —
            # not the taluks that sit inside them.
            normalized = normalize_district(d)
            key = normalized.strip().lower()
            if key and key not in seen_districts:
                seen_districts[key] = normalized
        district_list = sorted(seen_districts.values())

        # --------------------------------------------------------------
        # FIX: Year dropdown was previously missing entirely — the
        # template loops over `year_list`, but the view never built or
        # returned it, so only the static "All Years" option ever showed.
        # We derive distinct years straight from FarmVisitReport.visit_date
        # (auto-stamped when an executive logs a visit — there is no
        # manual date field in the form), newest year first.
        # --------------------------------------------------------------
        raw_years = (
            FarmVisitReport.objects.exclude(visit_date__isnull=True)
            .annotate(y=TruncYear("visit_date"))
            .values_list("y", flat=True)
            .distinct()
            .order_by("-y")
        )
        year_list = [y.year for y in raw_years if y]

        # --------------------------------------------------------------
        # FIX: Executive dropdown was previously built from ALL active
        # users (User.objects.filter(is_active=True)), which included
        # admin/staff accounts like "my_admin" alongside real field
        # executives. We now explicitly exclude staff and superuser
        # accounts so only genuine executives appear in the filter.
        # --------------------------------------------------------------
        executive_list = list(
            User.objects.filter(
                is_active=True,
                is_staff=False,
                is_superuser=False,
            )
            .values_list("username", flat=True)
            .distinct()
        )

        # Update context dictionary with calculated values
        context.update(
            {
                "total_revenue": total_rev,
                "total_visits": v_count,
                "active_executives": active_executives,
                "total_farms": total_farms_count,
                "total_sales_volume": vol_sold,
                "paid_orders_count": paid_orders_count,
                "avg_order_value": avg_order_value,
                "conversion_rate": conversion_rate,
                "hot_pct": hot_pct,
                "warm_pct": warm_pct,
                "cold_pct": cold_pct,
                "poultry_pct": poultry_pct,
                "aqua_pct": aqua_pct,
                "month_wise_labels_js": json.dumps(
                    month_wise_labels, cls=DjangoJSONEncoder
                ),
                "month_wise_data_js": json.dumps(
                    month_wise_data, cls=DjangoJSONEncoder
                ),
                "year_wise_labels_js": json.dumps(
                    year_wise_labels, cls=DjangoJSONEncoder
                ),
                "year_wise_data_js": json.dumps(
                    year_wise_data, cls=DjangoJSONEncoder
                ),
                "exec_labels_js": json.dumps(
                    exec_labels, cls=DjangoJSONEncoder
                ),
                "exec_revenue_js": json.dumps(
                    exec_revenue, cls=DjangoJSONEncoder
                ),
                "exec_conv_pct_js": json.dumps(
                    exec_conv_pct, cls=DjangoJSONEncoder
                ),
                "top_prod_labels_js": json.dumps(
                    top_prod_labels, cls=DjangoJSONEncoder
                ),
                "top_prod_revenue_js": json.dumps(
                    top_prod_revenue, cls=DjangoJSONEncoder
                ),
                "top_prod_qty_js": json.dumps(
                    top_prod_qty, cls=DjangoJSONEncoder
                ),
                "state_labels_js": json.dumps(
                    state_labels, cls=DjangoJSONEncoder
                ),
                "state_data_js": json.dumps(
                    state_data, cls=DjangoJSONEncoder
                ),
                # NOTE: these two are intentionally NOT pre-dumped with
                # json.dumps() like the other *_js keys above/below. The
                # dashboard.html template renders them through Django's
                # {{ chart_labels_js|json_script:"..." }} filter, which
                # already calls json.dumps() internally. Pre-dumping here
                # AND running it through json_script double-encodes the
                # list into a JSON string-of-a-string; JSON.parse() on the
                # JS side then only unwraps the outer layer and hands back
                # a plain string, which Chart.js iterates character-by-
                # character -- producing one bar per letter instead of
                # one bar per district. Keep these as raw Python lists.
                "chart_labels_js": chart_labels,
                "chart_counts_js": chart_counts,
                # Same rule applies here — dashboard.html renders this
                # through {{ farm_locations_js|json_script:"..." }}, which
                # already serializes it. Keep it as a raw Python list of
                # dicts, NOT json.dumps()'d, or the map JS will try to
                # JSON.parse() an already-stringified value and fail.
                "farm_locations_js": farm_locations_data,
                "stale_farms": stale_farms_table,
                "stale_farms_count": stale_farms_count,
                "prob_labels_js": json.dumps(
                    prob_labels, cls=DjangoJSONEncoder
                ),
                "prob_data_js": json.dumps(prob_data, cls=DjangoJSONEncoder),
                "bird_counts_js": json.dumps(
                    bird_counts, cls=DjangoJSONEncoder
                ),
                "pipeline_spread": pipeline_spread_agg,
                "top_farms": top_farms_table,
                "recent_visits": recent_visits,
                "state_list": state_list,
                "district_list": district_list,
                "executive_list": executive_list,
                "year_list": year_list,
            }
        )

    except Exception as e:
        logger.error(
            f"Error executing get_dashboard_context: {e}\n{traceback.format_exc()}"
        )

    return context

@staff_required
def dashboard_view(request):
    """View handler that renders the context into the HTML template."""
    context = get_dashboard_context(request)
    return render(request, 'dashboard.html', context)


@staff_required
def dashboard_home(request):
    context = get_dashboard_context(request)
    return render(request, 'crm_core/dashboard.html', context)


@staff_required
def dashboard_analytics(request):
    context = get_dashboard_context(request)
    return render(request, 'crm_core/analytics_report.html', context)


@staff_required
def executive_analytics_view(request):
    context = get_dashboard_context(request)
    return render(request, 'crm_core/analytics_report.html', context)

@staff_required
def clear_dashboard_data(request):
    """
    Clears all farm visit, product detail, and farm records from the CRM.
    Restricted to superusers/staff.
    """
    if request.method == 'POST':
        try:
            with transaction.atomic():
                VisitedProductDetail.objects.all().delete()
                FarmVisitReport.objects.all().delete()
                Farm.objects.all().delete()
            messages.success(request, "Dashboard data cleared successfully!")
        except Exception as e:
            logger.error(f"Failed to clear dashboard data: {str(e)}")
            messages.error(request, f"Error clearing data: {str(e)}")
    return redirect('dashboard_home')


# Place these at the end of backend/crm_core/views.py


@login_required(login_url='/crm/login/')
def get_location_details(request):
    """
    API endpoint for reverse geocoding latitude and longitude into location details.
    """
    lat = request.GET.get('lat')
    lon = request.GET.get('lon')

    if not lat or not lon:
        return JsonResponse({'error': 'Latitude and longitude required.'}, status=400)

    try:
        url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={lat}&lon={lon}"
        headers = {'User-Agent': 'AgriNutritionCRM/1.0'}
        res = requests.get(url, headers=headers, timeout=5)

        if res.status_code == 200:
            data = res.json()
            address = data.get('address', {})
            raw_district = address.get('state_district') or address.get('county') or address.get('district', '')
            return JsonResponse({
                'state': address.get('state', ''),
                'district': normalize_district(raw_district),
                'area': address.get('suburb') or address.get('village') or address.get('town') or address.get('city', '')
            })
        return JsonResponse({'error': 'Failed to fetch location data.'}, status=500)
    except Exception as e:
        logger.error(f"Geocoding error: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


# ==========================================
# 📲 WHATSAPP VISIT NOTIFICATION (LOCAL-ONLY)
# ==========================================
#
# WHY THIS IMPORTS SELENIUM/PYPERCLIP *INSIDE* THE VIEW, NOT AT THE
# TOP OF THIS FILE:
# Render has no Chrome binary and no display, so `selenium` and
# `pyperclip` are deliberately left out of the deployed requirements.
# If those imports sat at module level, Django would raise
# ImportError the moment this file loads — which would 500 every
# single view in the app on Render, not just this one. Importing
# them only when notify_farm_visit actually runs means:
#   - Render boots fine even without those packages installed.
#   - Locally (where you DO have selenium/pyperclip installed and a
#     real Chrome + WhatsApp Web session), the import succeeds and
#     the notify button works as expected.
#   - If someone ever hits this route on Render by mistake, they get
#     a clean 500 with a clear "WhatsApp automation isn't available
#     on this server" message instead of the whole app going down.
@staff_required
@csrf_exempt
def notify_farm_visit(request, visit_id):
    """
    Sends the given FarmVisitReport as a WhatsApp message to the
    executive's routed group, via a local Selenium-driven WhatsApp
    Web session. Only meaningful when run locally — see the module
    note above for why this can't work on Render.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required.'}, status=405)

    visit = get_object_or_404(
        FarmVisitReport.objects.select_related('farm', 'executive'),
        id=visit_id,
    )

    try:
        from .utils.whatsapp_routing import get_target_group
        from .utils.whatsapp_formatter import build_farm_visit_message
        from .utils.whatsapp_selenium import send_whatsapp_group_message
    except ImportError as e:
        logger.error(f"WhatsApp automation unavailable (missing dependency): {e}")
        return JsonResponse(
            {'error': 'WhatsApp automation is only available when run locally '
                      '(selenium/pyperclip not installed on this server).'},
            status=500,
        )

    try:
        group_title = get_target_group(visit)
    except ValueError as e:
        # Routing failures (no executive, no profile, no active group)
        # are expected/user-fixable — 400, not 500, and no browser is
        # launched at all.
        logger.warning(f"WhatsApp routing failed for visit {visit_id}: {e}")
        return JsonResponse({'error': str(e)}, status=400)

    try:
        message = build_farm_visit_message(visit)
        send_whatsapp_group_message(group_title, message)
    except Exception as e:
        logger.error(f"WhatsApp send failed for visit {visit_id}: {e}", exc_info=True)
        return JsonResponse(
            {'error': f'Failed to send WhatsApp message: {str(e)}'},
            status=500,
        )

    return JsonResponse({'success': True, 'group': group_title, 'visit_id': visit_id})
